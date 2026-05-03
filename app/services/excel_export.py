"""Build a styled Excel workbook from batch ADMET comparison rows."""

from __future__ import annotations

import io
from datetime import datetime, timezone

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _slug_tool(name: str) -> str:
    return {"SwissADME": "Swiss", "ADMETlab 3.0": "ADMETlab", "pkCSM": "pkCSM"}.get(
        name, name.replace(" ", "_")
    )


def _slug_metric(metric: str) -> str:
    return {
        "Molecular weight": "MW",
        "logP": "logP",
        "TPSA (Å²)": "TPSA",
        "Aqueous solubility (tool-specific)": "Solubility",
        "GI absorption (Swiss) / HIA (ADMETlab) / Intestinal % (pkCSM)": "GI_HIA_Intestinal",
        "Caco-2 permeability (log Papp)": "Caco2_logPapp",
        "BBB permeability (pkCSM)": "BBB_pkCSM",
    }.get(metric, metric[:24])


def pipeline_result_to_row(
    compound_id: str,
    smiles: str,
    pipeline_result: dict,
) -> dict:
    """Flatten one run_pipeline() result to a single dict for DataFrame export."""
    row: dict = {
        "ID": compound_id,
        "SMILES": smiles,
    }
    err = pipeline_result.get("errors") or {}
    row["Error_SwissADME"] = err.get("swissadme") or ""
    row["Error_ADMETlab"] = err.get("admetlab") or ""
    row["Error_pkCSM"] = err.get("pkcsm") or ""

    norm = pipeline_result.get("normalized") or {}
    for metric, per_tool in norm.items():
        ms = _slug_metric(metric)
        for tool_name, val in per_tool.items():
            ts = _slug_tool(tool_name)
            row[f"{ms}_{ts}"] = val
    return row


def build_batch_excel(df: pd.DataFrame) -> bytes:
    """
    Write DataFrame to xlsx with header styling, freeze pane, borders, column widths.
    Adds a short legend sheet.
    """
    legend_rows = [
        ("Spalte", "Bedeutung"),
        ("ID", "Bezeichner aus Datei oder fortlaufend"),
        ("SMILES", "Eingabe-SMILES"),
        ("Error_*", "Fehlermeldung falls Tool für diese Zeile fehlschlug"),
        ("MW_*", "Molekulargewicht je Tool"),
        ("logP_*", "lipophil (tool-spezifisch)"),
        ("TPSA_*", "polare Oberfläche; pkCSM: Surface Area Hinweis im Wert"),
        ("Solubility_*", "Löslichkeit — Skalen zwischen Tools nicht vergleichbar"),
        ("GI_HIA_Intestinal_*", "Resorption: Swiss Text / ADMET HIA / pkCSM %"),
        ("Caco2_*", "Caco-2 (log Papp) wo vorhanden"),
        ("BBB_pkCSM_*", "nur pkCSM"),
        ("", ""),
        ("Hinweis", "Alle Werte in silico; Laufzeit Batch = N × (3 Web-Tools)."),
    ]
    df_leg = pd.DataFrame(legend_rows[1:], columns=legend_rows[0])

    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Ergebnisse", index=False)
        df_leg.to_excel(writer, sheet_name="Legende", index=False)
        ws = writer.sheets["Ergebnisse"]
        _style_results_sheet(ws, df.shape[1])
        ws_leg = writer.sheets["Legende"]
        _style_legend_sheet(ws_leg)

    bio.seek(0)
    return bio.getvalue()


def _style_results_sheet(ws, ncols: int) -> None:
    header_fill = PatternFill("solid", fgColor="FF4472C4")
    header_font = Font(bold=True, color="FFFFFFFF", size=11)
    body_font = Font(size=10)
    thin = Side(style="thin", color="FFAAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_wrap = Alignment(wrap_text=True, vertical="top")
    align_header = Alignment(wrap_text=True, vertical="center", horizontal="center")

    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_header
        cell.border = border

    for row in range(2, ws.max_row + 1):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = body_font
            cell.alignment = align_wrap
            cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"

    widths = {"A": 14, "B": 52}
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        if letter in widths:
            ws.column_dimensions[letter].width = widths[letter]
        else:
            ws.column_dimensions[letter].width = 16


def _style_legend_sheet(ws) -> None:
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70
    for row in range(1, ws.max_row + 1):
        for col in range(1, 3):
            ws.cell(row=row, column=col).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=2).font = Font(bold=True)


def default_batch_filename() -> str:
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    return f"admet_batch_{ts}.xlsx"
